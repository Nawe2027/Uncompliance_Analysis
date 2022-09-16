# -*- coding: utf-8 -*-
"""
Created on Sat Aug  1 23:07:18 2020

@author: Administrator
"""
#%%

import pandas as pd 
import numpy as np 
from os import environ
import glob
import datetime as dt
from pandas import ExcelFile,DataFrame,to_numeric,to_datetime,DateOffset,read_csv,read_excel,concat,isnull,Series
from tkinter import Tk,Label,Button,StringVar,messagebox,Menu,OptionMenu,filedialog,HORIZONTAL
import dateutil.relativedelta
from time import sleep,strptime
from openpyxl import load_workbook,Workbook,drawing
from openpyxl.styles import Color, PatternFill,colors,Font,Alignment
from datetime import datetime,timedelta,date


MonthAnalysis = datetime.now()
MonthAnalysis = (MonthAnalysis.replace(day=1))
MonthAnalysis = MonthAnalysis.strftime("%b")
YearAnalysis = (datetime.now())
YearAnalysis = (YearAnalysis.replace(day=1)) + dateutil.relativedelta.relativedelta(days=-1)
YearAnalysis = int(YearAnalysis.strftime("%Y"))

def MasterFile():
    
    user = environ.get('USERNAME')
    
    df_GMC = read_excel("C:\\Users\\" + user + "\\Desktop\\Stealth\\Stealth Master File.xlsx",sheet_name="GMC List")
    Mds = df_GMC['GMC List'].drop_duplicates().tolist()
    
    df_CountryList = read_excel("C:\\Users\\" + user + "\\Desktop\\Stealth\\Stealth Master File.xlsx",sheet_name="CountryList")
    MarketUnitDic = df_CountryList.set_index('Country Name')['Market Unit Description'].to_dict()
    
    df_Calendar = read_excel("C:\\Users\\" + user + "\\Desktop\\Stealth\\Stealth Master File.xlsx",sheet_name="Calendar")
    
    MonthAnalysis = datetime.now()
    MonthAnalysis = (MonthAnalysis.replace(day=1))
    MonthAnalysis = MonthAnalysis.strftime("%b")
    
    df_Calendar = df_Calendar[['Country',MonthAnalysis]]
    
    
    df_Calendar = df_Calendar[~pd.isnull(df_Calendar[MonthAnalysis])] 

    
    ElegibleEx = df_Calendar['Country'].drop_duplicates().tolist()
    
    df_Exceptions = read_excel("C:\\Users\\" + user + "\\Desktop\\Stealth\\Stealth Master File.xlsx",sheet_name="Exceptions")
    
    df_Offenders = read_excel("C:\\Users\\" + user + "\\Desktop\\Stealth\\Stealth Master File.xlsx",sheet_name="Offender Tracker")
    
    return Mds,MarketUnitDic,ElegibleEx,df_Exceptions,df_Offenders


#%%
def ReportDates():
    starting = (datetime.now().replace(day=16) if datetime.now().day > 15 else datetime.now().replace(day=1)) - DateOffset(years=1)
    ending = datetime.now().replace(day=15) if datetime.now().day > 15 else datetime.now().replace(day=1) - timedelta(days=1)
    return starting,ending
#%%
def mydates(Year,Meses):
    from datetime import datetime
    MyTime = str(datetime.now())
    MyTime = MyTime.replace(":","_")
    Mes = strptime(Meses,'%b').tm_mon
    
    if datetime.now().day > 15:
        ReportEnd = date(Year, Mes, 1) + dateutil.relativedelta.relativedelta(days=14)
    else:
        ReportEnd = date(Year, Mes, 1) + dateutil.relativedelta.relativedelta(days=-1)
   
    ReportEnd_MyTE = ReportEnd.strftime('%m/%d/%Y')
    
    ReportStart = (date(Year, Mes, 1)) + dateutil.relativedelta.relativedelta(months=-12)
    
    if datetime.now().day > 15:
        ReportStart = (date(Year, Mes, 16)) + dateutil.relativedelta.relativedelta(months=-12)
    else:
        ReportStart = (date(Year, Mes, 1)) + dateutil.relativedelta.relativedelta(months=-12)    
    
    #BORRAR ESTO
    ReportEnd = date(Year, Mes, 1) + dateutil.relativedelta.relativedelta(days=-1)
    ReportStart = (date(Year, Mes, 1)) + dateutil.relativedelta.relativedelta(months=-12)    

    ReportStart_MyTE = (date(Year -1, 1, 1))
    ReportStart_MyTE = ReportStart_MyTE.strftime('%m/%d/%Y')
    
    CYStart = date(Year, 1, 1)
    CYEnd = date(Year, 12, 31)
    Rolling_6 = (ReportEnd) + dateutil.relativedelta.relativedelta(months=-6)
    Rolling_3 = (ReportEnd) + dateutil.relativedelta.relativedelta(months=-3)
    Rolling_2 = (ReportEnd) + dateutil.relativedelta.relativedelta(months=-2)
    
    return ReportEnd,ReportEnd_MyTE,ReportStart,ReportStart_MyTE,CYStart,CYEnd,Rolling_6,Rolling_3,Rolling_2

#%%    
def ExpenseAnalysis(ElegibleEx):
    
    user = environ.get('USERNAME')
    
    IsoCodes = pd.read_excel("C:\\Users\\" + user + "\\Desktop\\Stealth\\Country Codes (1).xlsx")
    path = "C:\\Users\\" + user + "\\Desktop\\Stealth" # use your path
    all_files = glob.glob(path + "//Expense Details*.csv")
    
    df = None 
    for filename in all_files:
        if df is None:
            df = pd.read_csv(filename)
            df = df[['EnterpriseId',	'PersonnelID','EmploymentStatusDesc','EmployeeCountry',	'CareerLvlGrpDesc',	'ParentWBS', 'ProcessedDttm' ,	'textbox27',	'ChargeType',	'ExpenseType',	'ExpenseFromDate',	'ExpenseToDate',	'NumberOfDays',	'SplitPercent',	'ReceiptAmount',	'CountryOfExpense',	'AdditionalInformation1','CompanyCode','CareerLvlGrpDesc']]
#            df = df[(df['EmploymentStatusDesc'] == "Active") | (df['EmploymentStatusDesc'] == "Inactive")]
        else:
            df1 = pd.read_csv(filename)
            df1 = df1[['EnterpriseId',	'PersonnelID','EmploymentStatusDesc',	'EmployeeCountry',	'CareerLvlGrpDesc',	'ParentWBS', 'ProcessedDttm' ,	'textbox27',	'ChargeType',	'ExpenseType',	'ExpenseFromDate',	'ExpenseToDate',	'NumberOfDays',	'SplitPercent',	'ReceiptAmount',	'CountryOfExpense',	'AdditionalInformation1','CompanyCode','CareerLvlGrpDesc']]
#            df1 = df1[(df1['EmploymentStatusDesc'] == "Active") | (df1['EmploymentStatusDesc'] == "Inactive")]
            df = df.append(df1)
    del(df1)
    
    df.rename(columns={'CareerLvlGrpDesc':'CareerLevel'},inplace=True)

    Ex_Cl = df[['EnterpriseId','CareerLevel']].drop_duplicates()
    Ex_Cl = Ex_Cl.set_index('EnterpriseId')['CareerLevel'].to_dict()

    Ex_St = df[['EnterpriseId','EmploymentStatusDesc']].drop_duplicates()
    Ex_St = Ex_St.set_index('EnterpriseId')['EmploymentStatusDesc'].to_dict()

    df['CountryOfExpense'] = df['CountryOfExpense'].str.strip()
    df['EmployeeCountry'] = df['EmployeeCountry'].str.strip()
    df['ExpenseType'] = df['ExpenseType'].str.strip()
    df['AdditionalInformation1'] = df['AdditionalInformation1'].str.strip()

    df.sort_values(['ProcessedDttm','ReceiptAmount','SplitPercent'],ascending=[False,False,False],inplace=True)           
    df.drop_duplicates(['textbox27'],inplace=True)   
    df['TE'] = df["AdditionalInformation1"].str[-2:]
    df=IsoCodes[['TE','CountryName']].merge(df,on='TE',how='right')

    df.loc[(df['ExpenseType'] == "Per Diem - International"),['CountryOfExpense']] = df["CountryName"]

    df = ConvertName(df,"CountryOfExpense","EmployeeCountry")

    df['CountryOfExpense'] = df['CountryOfExpense'].str.strip()

    df['EmployeeCountry'] = df['EmployeeCountry'].str.strip()
            
    df = df[(df['EmployeeCountry'] != df['CountryOfExpense'])]
    
    df = df[(pd.isnull(df['CountryOfExpense'])==False)]
    
    
    df.loc[(pd.isnull(df['ExpenseToDate'])),['ExpenseToDate']] = df["ExpenseFromDate"]
    df = df.drop_duplicates(keep=False)
    df  = df[['EnterpriseId','EmployeeCountry','CountryOfExpense','ParentWBS','ExpenseType','ExpenseToDate','ExpenseFromDate','CompanyCode']]
    
    df['ExpenseFromDate'] = pd.to_datetime(df['ExpenseFromDate'])
    df['ExpenseToDate'] = pd.to_datetime(df['ExpenseToDate'])
    df['DateVal'] = [pd.date_range(s, e, freq='d') for s, e in
                  zip(pd.to_datetime(df['ExpenseFromDate']), pd.to_datetime(df['ExpenseToDate']))]

    df = df.explode('DateVal').drop(['ExpenseFromDate', 'ExpenseToDate'], axis=1)
    df = ConvertName(df,"CountryOfExpense","EmployeeCountry")
    
    df = df[df['CountryOfExpense'].isin(ElegibleEx)]

    df['EE+D'] = df['EnterpriseId'] + df['CountryOfExpense']
    df.drop_duplicates(['EnterpriseId','CountryOfExpense','DateVal'],inplace=True)   
    df['Year'], df['Month'] = df['DateVal'].dt.year, df['DateVal'].dt.month_name()
    df['Period'] = df['Month'] + '-' + df['Year'].astype(str)
    
    df['CareerLevel'] = df['EnterpriseId'].map(Ex_Cl)
    
    df = df[df['DateVal'] > '2019-12-16']

    df_PP = DataFrame.pivot_table(df,values='DateVal',
              index=['EnterpriseId','CountryOfExpense'],columns=['Period'],aggfunc=lambda x: len(x.unique())).fillna(0)
    return df_PP,df,Ex_Cl,Ex_St
#%%

def OrdenarColumnas(Starting,df_PP,df_EX):
    colDate = Starting.replace(day=1)
    colName = []
    for i in range(12):
        colName.append(colDate.strftime('%B-%Y'))
        if colDate.strftime('%B-%Y') not in df_PP: df_PP[colDate.strftime('%B-%Y')] = 0
        if colDate.month==12: colDate = colDate.replace(month=1,year=colDate.year+1)
        else: colDate = (colDate - timedelta(days=1)).replace(day=1,month=colDate.month+1,year=colDate.year)
    df_EX = df_EX[colName]; df_PP = df_PP[colName]
    return df_EX,df_PP,colName

#%%
    
def ConvertName(df,Destination,Home):
    df[Destination] = df[Destination].str.strip()
    df.loc[(df[Destination] == "US"),[Destination]] = "USA"
    df.loc[(df[Destination] == "UK"),[Destination]] = "United Kingdom"
    df.loc[(df[Destination] == "Costa Rica USD"),[Destination]] = "Costa Rica"
    df.loc[(df[Destination] == "Costa Rica CRC"),[Destination]] = "Costa Rica"
    df.loc[(df[Destination] == "China"),[Destination]] = "China/Mainland"
    df.loc[(df[Destination] == "Hong Kong"),[Destination]] = "China/Hong Kong SAR"
    df.loc[(df[Destination] == "China/Hong Kong"),[Destination]] = "China/Hong Kong SAR"
    df.loc[(df[Destination] == "Macau"),[Destination]] = "China/Macao SAR"
    df.loc[(df[Destination] == "Taiwan"),[Destination]] = "China/Taiwan"   
    df.loc[(df[Destination] == "Cote D'Ivoire"),[Destination]] = "Côte d'Ivoire"
    df.loc[(df[Destination] == "Còte D'Ivoire"),[Destination]] = "Côte d'Ivoire"    
    df.loc[(df[Destination] == "Côte d'Ivoire"),[Destination]] = "Côte d'Ivoire"      
    df.loc[(df[Destination] == "Russian Fed."),[Destination]] = "Russian Federation"
    df.loc[(df[Destination] == "Russia"),[Destination]] = "Russian Federation"
    df.loc[(df[Destination] == "Korea"),[Destination]] = "South Korea"
    df.loc[(df[Destination] == "Utd.Arab Emir."),[Destination]] = "United Arab Emirates"
    df.loc[(df[Destination] == "Slovak Republic"),[Destination]] = "Slovakia"
    df.loc[(df[Destination] == "UAE"),[Destination]] = "United Arab Emirates"
    
    df[Home] = df[Home].str.strip()
    df.loc[(df[Home] == "US"),[Home]] = "USA"
    df.loc[(df[Home] == "UK"),[Home]] = "United Kingdom"
    df.loc[(df[Home] == "Costa Rica USD"),[Home]] = "Costa Rica"
    df.loc[(df[Home] == "Costa Rica CRC"),[Home]] = "Costa Rica"
    df.loc[(df[Home] == "China"),[Home]] = "China/Mainland"
    df.loc[(df[Home] == "Hong Kong"),[Home]] = "China/Hong Kong SAR"
    df.loc[(df[Home] == "China/Hong Kong"),[Home]] = "China/Hong Kong SAR"    
    df.loc[(df[Home] == "Macau"),[Home]] = "China/Macao SAR"
    df.loc[(df[Home] == "Taiwan"),[Home]] = "China/Taiwan"   
    df.loc[(df[Home] == "Cote D'Ivoire"),[Home]] = "Côte d'Ivoire"
    df.loc[(df[Home] == "Còte D'Ivoire"),[Home]] = "Côte d'Ivoire"    
    df.loc[(df[Home] == "Côte d'Ivoire"),[Home]] = "Côte d'Ivoire"    
    df.loc[(df[Home] == "Russian Fed."),[Home]] = "Russian Federation"
    df.loc[(df[Home] == "Russia"),[Home]] = "Russian Federation"
    df.loc[(df[Home] == "Korea"),[Home]] = "South Korea"
    df.loc[(df[Home] == "Utd.Arab Emir."),[Home]] = "United Arab Emirates"
    df.loc[(df[Home] == "UAE"),[Home]] = "United Arab Emirates"
    df.loc[(df[Home] == "Slovak Republic"),[Home]] = "Slovakia"
    return df

#%%
    
def ImportStealthSSRS():
    
    user = environ.get('USERNAME')

    df_TP = read_csv("C:\\Users\\" + user + "\\Desktop\\Stealth\\Stealth Report.csv")
#    df_TP.columns = df_TP.iloc[2]
#    df_TP = df_TP.iloc[3:]
    df_TP = df_TP[(df_TP['Travel_Plan_Number'] != "TRA0455419")]
    df_TP = df_TP[(df_TP['Travel_Plan_Number'] != "TRA0474094")]
    df_TP['Travel_End_Date'] = to_datetime(df_TP['Travel_End_Date'])
    df_TP['Travel_Start_Date'] = to_datetime(df_TP['Travel_Start_Date'])
    df_TP.sort_values(['Travel_End_Date'],ascending=[False],inplace=True)           
    df_TP.drop_duplicates(['Enterprise_ID','Destination_Country'],inplace=True)   
    df_TP['Policy_Type']  = df_TP['Policy_Type'].astype(str)
#    PolicyOtherCategories = ['Remote Work','Non-Accenture Policy','Policy 1085 Exception']
    df_TP = df_TP[(df_TP['Policy_Type'] == "710") | (df_TP['Policy_Type'] == "810") | (df_TP['Policy_Type'] == "1430") | (df_TP['Policy_Type'] == "740") | (df_TP['Policy_Type'] == "750") | (df_TP['Policy_Type'] == "Other") ] #DEJAMOS SOLO POLICY ELEGIBLES
    df_TP['EnterpriseId'] = df_TP['Enterprise_ID']
    df_TP['Location1'] = df_TP['Destination_Country']
    df_TP['Code'] = df_TP['Enterprise_ID'] + df_TP['Destination_Country']
    df_TP['SD After Merge'] = df_TP['Travel_Start_Date'].groupby(df_TP['Code']).transform('min')
    df_TP['ED After Merge'] = df_TP['Travel_End_Date'].groupby(df_TP['Code']).transform('max')    
    return df_TP

#%%i 
        
def ImportLBD(ElegibleEx):
    user = environ.get('USERNAME')

    path = "C:\\Users\\" + user + "\\Desktop\\Stealth" # use your path
    all_files = glob.glob(path + "//LBD*.csv")
    li = []
    
    for filename in all_files:
        df = read_csv(filename, index_col=None, header=0)
        df = df[(df['EmploymentStatusDesc'] == "Active") | (df['EmploymentStatusDesc'] == "Inactive")]
        li.append(df)
        
    df_LBD = concat(li, axis=0, ignore_index=True)

#    pnmr = df_LBD['EnterpriseId'].drop_duplicates().tolist()
#    pnmr = df_LBD[''].drop_duplicates().tolist()
#
#
#    list_EID = df_LBD['PersonnelNbr'].drop_duplicates().tolist() #PersonnelNbr
#    with open('C:\\Users\\Administrator\\Desktop\\List.txt', 'w') as f:
#        for item in list_EID:
#            f.write("%s\n" % item)
 
    
#    df1 = pd.read_excel('C:\\Users\\Administrator\\Downloads\\EXPORT1.xlsx')
#    df_LBD = df1.merge(df_LBD,on='EnterpriseId',how='right')
    df_LBD.rename(columns={'HostCountry':'Location1'},inplace=True)
    df_LBD = ConvertName(df_LBD,"Location1","HomeCountry")
    df_LBD = df_LBD[(df_LBD['Location1'].isin(ElegibleEx))]
    df_LBD.rename(columns={'Name of Company Code':'Name_of_Company_Code'},inplace=True)
    df_LBD.rename(columns={'EmployeeStatus':'EmploymentStatusDesc'},inplace=True)
    
    df_LBD['Date'] = pd.to_datetime(df_LBD['Date'])
    df_LBD['Month'] = pd.to_datetime(df_LBD['Date'], format='%m').dt.month_name()
    df_LBD['Year'] = pd.DatetimeIndex(df_LBD['Date']).year
    df_LBD['Period'] = df_LBD['Month'] + '-' + df_LBD['Year'].astype(str)
    df_LBD.rename(columns={'Date':'DateVal'},inplace=True)
 
    df_LBD.rename(columns={'SubmitEnterpriseId':'EnterpriseId'},inplace=True)
    df_LBD.rename(columns={'Career Level Group':'CareerLevel'},inplace=True)

    df_LBD = df_LBD[['EnterpriseId','EmploymentStatusDesc','HomeCountry','Location1','TypeOfDay','DateVal','Period','CompanyCd','Name_of_Company_Code','WBS_Element','CareerLevel']]
    df_LBD.drop_duplicates(inplace=True)
    df_LBD.reset_index(drop=True,inplace=True)
    TE_St = df_LBD[['EnterpriseId','EmploymentStatusDesc']].drop_duplicates()
    TE_St = TE_St.set_index('EnterpriseId')['EmploymentStatusDesc'].to_dict()
    Te_Cl = df_LBD[['EnterpriseId','CareerLevel']].drop_duplicates()
    Te_Cl = Te_Cl.set_index('EnterpriseId')['CareerLevel'].to_dict()
    df_LBD = df_LBD[['EnterpriseId','HomeCountry','Location1','TypeOfDay','DateVal','Period','CompanyCd','Name_of_Company_Code','WBS_Element','CareerLevel']]
    df_PP = DataFrame.pivot_table(df_LBD,values='DateVal',
              index=['EnterpriseId','Location1'],columns=['Period'],aggfunc=lambda x: len(x.unique())).fillna(0)
    df_LBD['Code'] = df_LBD['EnterpriseId'] + df_LBD['Location1']
    return df_LBD,df_PP,Te_Cl,TE_St

#%%

def analisisDOA_MyTE(df_LBD,df_TP):
    df_LBD['Enterprise ID'] = df_LBD['EnterpriseId']
    df_LBD['Code'] = df_LBD['EnterpriseId'] + df_LBD['Location1']
    TravelPlans = df_TP['Code'].drop_duplicates().tolist()
    df_LBD = df_LBD[(df_LBD['Code'].isin(TravelPlans))]
    df = df_LBD[['Code','DateVal']].merge(df_TP[['Code','Enterprise_ID','Destination_Country','Travel_Plan_Number','SD After Merge','ED After Merge']],on='Code',how='right')
    df = df.drop_duplicates(['Code','DateVal'])
    df['DOA'] = None
    df.loc[(df['DateVal']<=df['ED After Merge']), 'DOA'] = 'DAYS IN'
    df['DOA'].fillna('DAYS OUT',inplace=True)
    #df = df[df['DOA']=='DAYS IN'][['Code','DateVal','DOA']].drop_duplicates()
    df = df.drop_duplicates(['Code','DateVal'])
    df = df[df['DOA']=='DAYS OUT']
    df = df[pd.isnull(df['DateVal'])==False]
    df_PP_DOA = DataFrame.pivot_table(df[df['DOA']=='DAYS OUT'],
                                  values='DateVal' ,index=['Code'],aggfunc=lambda x: len(x.unique())).fillna(0)
    df_PP_DOA.rename(columns={'DateVal':'Days Out of Assignment (total number of days out)'},inplace=True)
    del(df_LBD)
    return df_PP_DOA

#%%

def analisisDOA_Expense(df_EX,df_TP):
    df_EX['Enterprise ID'] = df_EX['EnterpriseId']
    df_EX['Code'] = df_EX['EnterpriseId'] + df_EX['CountryOfExpense']
    TravelPlans = df_TP['Code'].drop_duplicates().tolist()
    df_EX = df_EX[(df_EX['Code'].isin(TravelPlans))]
    df = df_EX[['Code','DateVal']].merge(df_TP[['Code','Enterprise_ID','Destination_Country','Travel_Plan_Number','SD After Merge','ED After Merge']],on='Code',how='right')
    df = df.drop_duplicates(['Code','DateVal'])
    df['DOA'] = None
    df.loc[(df['DateVal']<=df['ED After Merge']), 'DOA'] = 'DAYS IN'
    df['DOA'].fillna('DAYS OUT',inplace=True)
    #df = df[df['DOA']=='DAYS IN'][['Code','DateVal','DOA']].drop_duplicates()
    df = df.drop_duplicates(['Code','DateVal'])
    df = df[df['DOA']=='DAYS OUT']
    df = df[pd.isnull(df['DateVal'])==False]
    df_PP_DOA = DataFrame.pivot_table(df[df['DOA']=='DAYS OUT'],
                                  values='DateVal' ,index=['Code'],aggfunc=lambda x: len(x.unique())).fillna(0)
    df_PP_DOA.rename(columns={'DateVal':'Expense Days Out of Assignment'},inplace=True)
    return df_PP_DOA

#%%

def PopulationDOA(DOA_MyTE,DOA_Expense):
    Population = DOA_MyTE.join(DOA_Expense, how='outer')
    Population.fillna(0,inplace=True)
    Population["Total Days Out (MyTE vs Expense)"] = Population[["Days Out of Assignment (total number of days out)", "Expense Days Out of Assignment"]].max(axis=1)
    del globals()["DOA_MyTE"]
    del globals()["DOA_Expense"]
    return Population

#%%
def DOA_In30(df_TP,df_LBD,df_ExPP):
    df_TP['ED'] = df_TP['ED After Merge']
    df_TP['ED After Merge'] =  df_TP['ED After Merge'] + timedelta(days=30)
    MyTE = analisisDOA_MyTE(df_LBD,df_TP)
    Expense = analisisDOA_Expense(df_ExPP,df_TP)
    Population = MyTE.join(Expense, how='outer')
    Population.reset_index(inplace=True)  
    DOA_Post = Population['Code'].drop_duplicates().tolist()
    return DOA_Post

#%%

def Selecting_Stealth_Population(df_EX,df_PP,df_TP,DOAStealth):
    df_EX.reset_index(inplace=True)  
    df_PP.reset_index(inplace=True)  
    df_EX['Code'] = df_EX['EnterpriseId'] + df_EX['CountryOfExpense']
    df_PP['Code'] = df_PP['EnterpriseId'] + df_PP['Location1']
    df_EX['Include'] = np.nan
    df_PP['Include'] = np.nan
    TravelPlans = df_TP['Code'].drop_duplicates().tolist()
    df_EX.loc[((df_EX['Code'].isin(TravelPlans))),['Include']] = "No"
    df_PP.loc[((df_PP['Code'].isin(TravelPlans))),['Include']] = "No"
    df_DOA.reset_index(inplace=True)
    Population_List = df_DOA['Code'].drop_duplicates().tolist()
    df_EX.loc[(df_EX['Code'].isin(Population_List)),['Include']] = "Include"
    df_PP.loc[(df_PP['Code'].isin(Population_List)),['Include']] = "Include"
    df_EX.loc[(df_EX['Code'].isin(DOAStealth)),['Include']] = "Include"
    df_PP.loc[(df_PP['Code'].isin(DOAStealth)),['Include']] = "Include"
    df_EX = df_EX[(df_EX['Include'] != "No")]
    df_PP = df_PP[(df_PP['Include'] != "No")]
    df_EX = df_EX.drop(columns=["EnterpriseId", "CountryOfExpense","Include"])
    df_PP = df_PP.drop(columns=["EnterpriseId", "Location1","Include"])
    df_EX.set_index('Code',inplace=True)
    df_PP.set_index('Code',inplace=True)
    df = pd.merge(df_PP, df_EX, left_index=True, right_index=True)
    return df,df_EX,df_PP

#%%

def MaxandSource(Starting,df_DayPivot,df_EX,df_PP):
    colDate = Starting.replace(day=1)
    colName1 = []
    for i in range(12):
        df_DayPivot[str(colDate.strftime('%B-%Y'))] = df_DayPivot[[str(colDate.strftime('%B-%Y')) + "_x", str(colDate.strftime('%B-%Y')) + "_y"]].max(axis=1)
        df_DayPivot["Source " + str(colDate.strftime('%B-%Y'))] = ""
        df_DayPivot.loc[(df_DayPivot[str(colDate.strftime('%B-%Y')) + "_x"] > df_DayPivot[str(colDate.strftime('%B-%Y')) + "_y"]),["Source " + str(colDate.strftime('%B-%Y'))]] = "Work Location"
        df_DayPivot.loc[(df_DayPivot[str(colDate.strftime('%B-%Y')) + "_y"] > df_DayPivot[str(colDate.strftime('%B-%Y')) + "_x"]),["Source " + str(colDate.strftime('%B-%Y'))]] = "Expenses"
        df_DayPivot.loc[(df_DayPivot[str(colDate.strftime('%B-%Y')) + "_y"] == df_DayPivot[str(colDate.strftime('%B-%Y')) + "_x"]),["Source " + str(colDate.strftime('%B-%Y'))]] = "Work Location/Expenses"
        df_DayPivot.loc[(df_DayPivot[str(colDate.strftime('%B-%Y'))] == 0),["Source " + str(colDate.strftime('%B-%Y'))]] = "N/A"
        colName1.append(str(colDate.strftime('%B-%Y')))
        colName1.append("Source " + str(colDate.strftime('%B-%Y')))
        if colDate.strftime('%B-%Y') not in df_PP: df_PP[colDate.strftime('%B-%Y')] = 0
        if colDate.month==12: colDate = colDate.replace(month=1,year=colDate.year+1)
        else: colDate = (colDate - timedelta(days=1)).replace(day=1,month=colDate.month+1,year=colDate.year)
    df_DayPivot = df_DayPivot[colName1]
    
    EX_Popu = df_EX.index.drop_duplicates().tolist()
    PP_Popu = df_PP.index.drop_duplicates().tolist()
    
    df_EX1 = df_EX[~df_EX.index.isin(PP_Popu)]
    df_PP1 = df_PP[~df_PP.index.isin(EX_Popu)]

    colDate = Starting.replace(day=1)
    for i in range(12):
        df_EX1["Source " + str(colDate.strftime('%B-%Y'))] = "Expenses"
        df_EX1.loc[(df_EX1[str(colDate.strftime('%B-%Y'))] == 0),["Source " + str(colDate.strftime('%B-%Y'))]] = "N/A"
        df_PP1["Source " + str(colDate.strftime('%B-%Y'))] = "Work Location"
        df_PP1.loc[(df_PP1[str(colDate.strftime('%B-%Y'))] == 0),["Source " + str(colDate.strftime('%B-%Y'))]] = "N/A"
    
        if colDate.strftime('%B-%Y') not in df_PP: df_PP[colDate.strftime('%B-%Y')] = 0
        if colDate.month==12: colDate = colDate.replace(month=1,year=colDate.year+1)
        else: colDate = (colDate - timedelta(days=1)).replace(day=1,month=colDate.month+1,year=colDate.year)
    
    df_EX1 = df_EX1[colName1]
    df_PP1 = df_PP1[colName1]
    df_DayPivot = df_DayPivot.append(df_EX1)
    df_DayPivot = df_DayPivot.append(df_PP1)
    del globals()["df_EX"]
    del globals()["df_PP"]
    return df_DayPivot,colName1

#%%
    



def Counting_Rolling_Calendar(df_DayPivot,colName,MonthAnalysis,YearAnalysis):
    ReportEnd,ReportEnd_MyTE,ReportStart,ReportStart_MyTE,CYStart,CYEnd,Rolling_6,Rolling_3,Rolling_2 = mydates(YearAnalysis,MonthAnalysis) #CON ESTA FUNCION DETERMINO LAS FECHAS

    if ReportEnd > (date(YearAnalysis, 1, 1)) and ReportEnd < (date(YearAnalysis, 4, 6)):
        FYReportStart = (date(YearAnalysis - 1, 4, 1))
    else:
        FYReportStart = (date(YearAnalysis, 4, 1))
        
    #CALULA LOS DIAS EN LOS ULTIMOS 12 MESES, SE USA CUANDO ES ROLLING Y POR EL MAXA NO LLEGA
    df_DayPivot['DAYS IN LAST 12'] = 0
#    for i in range(colName.index(ReportEnd.strftime('%B-%Y')),colName.index(ReportEnd.strftime('%B-%Y'))-12,-1): 
#        df_DayPivot['DAYS IN LAST 12'] += df_DayPivot[colName[i]]  
#             
    df_DayPivot['DAYS IN LAST 12'] = df_DayPivot[colName].sum(axis=1)
    #CALULA LOS DIAS EN LOS ULTIMOS 6 MESES, SE USA CUANDO ES CY Y POR EL MAXA NO LLEGA
    df_DayPivot['DAYS IN LAST 6'] = 0
    for i in range(colName.index(ReportEnd.strftime('%B-%Y')),colName.index(ReportEnd.strftime('%B-%Y'))-6,-1):
        print(colName[i])
        df_DayPivot['DAYS IN LAST 6'] += df_DayPivot[colName[i]]   
    #CALULA LOS DIAS EN LOS ULTIMOS 3 MESES, SE USA CUANDO ES 6  MONTH ROLLING Y POR EL MAXA NO LLEGA
    df_DayPivot['DAYS IN LAST 3'] = 0
    for i in range(colName.index(ReportEnd.strftime('%B-%Y')),colName.index(ReportEnd.strftime('%B-%Y'))-3,-1): 
        df_DayPivot['DAYS IN LAST 3'] += df_DayPivot[colName[i]]    
    #CALULA LOS DIAS EN LOS ULTIMOS 2 MESES, SE USA CUANDO ES 6  MONTH ROLLING Y POR EL MAXA NO LLEGA
    df_DayPivot['DAYS IN LAST 2'] = 0
    for i in range(colName.index(ReportEnd.strftime('%B-%Y')),colName.index(ReportEnd.strftime('%B-%Y'))-2,-1): 
        df_DayPivot['DAYS IN LAST 2'] += df_DayPivot[colName[i]]     
    #CALCULA LOS DIAS EN EL FY
    df_DayPivot['FY DAYS'] = 0
    for i in range(colName.index(FYReportStart.strftime('%B-%Y')),colName.index(ReportEnd.strftime('%B-%Y'))): 
        df_DayPivot['FY DAYS'] += df_DayPivot[colName[i]] 
    
    colDate = CYStart.replace(day=1)
    colName = []
    Mes = strptime(MonthAnalysis,'%b').tm_mon -1 
    
    for i in range(Mes):
        colName.append(colDate.strftime('%B-%Y'))
        if colDate.strftime('%B-%Y') not in df_DayPivot: df_DayPivot[colDate.strftime('%B-%Y')] = 0
        if colDate.month==12: colDate = colDate.replace(month=1,year=colDate.year+1)
        else: colDate = (colDate - timedelta(days=1)).replace(day=1,month=colDate.month+1,year=colDate.year)
    
    df_DayPivot['CY'] = df_DayPivot[colName].sum(axis=1)
    
    return df_DayPivot,colName

#%%
    
def Importing_Thresholds():
    
    user = environ.get('USERNAME')
        
    df_Threashold = read_excel("C:\\Users\\" + user + "\\Desktop\\Stealth\\Stealth Threshold.xlsx",sheet_name="New")
    df_Threashold = df_Threashold[['Citizenship','Host Country','Lowest Threshold (per trip)','Lowest Threshold (Cumm)','Months','Period','Action if exceeding Stealth thresholds']]
    df_Threashold.rename(columns={'Lowest Threshold (per trip)':'Single Trip Stealth Threshold'},inplace=True)
    df_Threashold.rename(columns={'Lowest Threshold (Cumm)':'Cumulative Stealth Threshold'},inplace=True)
    df_Threashold.rename(columns={'Action if exceeding Stealth thresholds':'ET&I Remediation Action'},inplace=True)
    df_Threashold['Period'] = df_Threashold['Period'].str.strip()
    return df_Threashold

#%%
    
def StraightDays_Fist(df_ExPP,df_LBD):
    df_ExPP['CodeThreshold'] = df_ExPP['EmployeeCountry'] + df_ExPP['CountryOfExpense']
    df_LBD['CodeThreshold'] = df_LBD['HomeCountry'] + df_LBD['Location1']
    df_ExPP = df_ExPP[['Code','DateVal','CodeThreshold']]
    df_LBD = df_LBD[['Code','DateVal','CodeThreshold']]
    df_StraightDays = df_LBD.append(df_ExPP, ignore_index=True)
    return df_StraightDays

#%%
def Straight_days(df,df_Threashold,MonthAnalysis,YearAnalysis):
    df['DateVal'] = pd.to_datetime(df['DateVal'])
    df.sort_values(['Code','DateVal'],ascending=[True,True],inplace=True)           
    df = df.drop_duplicates(['Code','DateVal'])
    df['CONSECUTIVE_DAY'] = False
    df.loc[(df['DateVal'] - df['DateVal'].shift(1))/np.timedelta64(1,'D') == 1.0, 'CONSECUTIVE_DAY'] = True
    df.loc[(df['DateVal'] - df['DateVal'].shift(1))/np.timedelta64(1,'D') == 1, 'CONSECUTIVE_DAY'] = True
    df['Date_Cut'] = df['DateVal'].where(df['CONSECUTIVE_DAY'] == False).groupby(df['Code']).transform('max')
    df['Include'] = "Include"
    df.loc[(df['DateVal'] < df['Date_Cut']),['Include']] = "Delete"
    df = df[(df['Include'] == "Include" )]
    df_Threashold['CodeThreshold'] = df_Threashold['Citizenship'] + df_Threashold['Host Country']
    df_Threashold.rename(columns={'Period':'Month Period'},inplace=True)
    df = df_Threashold[['CodeThreshold','Months','Month Period']].merge(df,on='CodeThreshold',how='right')
    ReportEnd,ReportEnd_MyTE,ReportStart,ReportStart_MyTE,CYStart,CYEnd,Rolling_6,Rolling_3,Rolling_2 = mydates(YearAnalysis,MonthAnalysis)
    
    if ReportEnd > (date(YearAnalysis, 1, 1)) and ReportEnd < (date(YearAnalysis, 4, 6)):
        FYReportStart = (date(YearAnalysis - 1, 4, 1))
    else:
        FYReportStart = (date(YearAnalysis, 4, 1))
        
    df['FY'] = FYReportStart;df['FY'] = pd.to_datetime(df['FY'])
    df['Rolling 2'] = Rolling_2;df['Rolling 2'] = pd.to_datetime(df['Rolling 2'])
    df['Rolling 3'] = Rolling_3;df['Rolling 3'] = pd.to_datetime(df['Rolling 3'])
    df['Rolling 6'] = Rolling_6;df['Rolling 6'] = pd.to_datetime(df['Rolling 6'])
    df['Rolling 12'] = ReportStart;df['Rolling 12'] = pd.to_datetime(df['Rolling 12'])
    df['CY'] = CYStart;df['CY'] = pd.to_datetime(df['CY'])
    df['Include'] = "Include"
    df.loc[(df['DateVal'] < df['Rolling 2']) & ((df['Month Period'] == "Rolling") | (df['Month Period'] == "rolling")) & (df['Months'] == 2),['Include']] = "Delete"
    df.loc[(df['DateVal'] < df['Rolling 3']) & ((df['Month Period'] == "Rolling") | (df['Month Period'] == "rolling")) & (df['Months'] == 3),['Include']] = "Delete"
    df.loc[(df['DateVal'] < df['Rolling 6']) & ((df['Month Period'] == "Rolling") | (df['Month Period'] == "rolling")) & (df['Months'] == 6),['Include']] = "Delete"
    df.loc[(df['DateVal'] < df['Rolling 12']) & ((df['Month Period'] == "Rolling") | (df['Month Period'] == "rolling")) & (df['Months'] == 12),['Include']] = "Delete"
    df.loc[(df['DateVal'] < df['CY']) & ((df['Month Period'] == "Calendar") | (df['Month Period'] == "calendar") ),['Include']] = "Delete"
    df.loc[(df['DateVal'] < df['FY']) & ((df['Month Period'] == "Tax Year") | (df['Month Period'] == "Tax year") ),['Include']] = "Delete"
    df = df[(df['Include'] == "Include" )]
    
    return df
#%%
def StraightDays_Dataframe(df_StraightDays):    
    df_StraightDays = df_StraightDays[['Code','DateVal','Include']]
    df_StraightDays = df_StraightDays.drop_duplicates(['Code','DateVal'])
    df_StraightDays['Total Single Trip'] = df_StraightDays['Include'].where(df_StraightDays['Include'] == "Include").groupby(df_StraightDays['Code']).transform('count')
    df_StraightDays = df_StraightDays[['Code','Total Single Trip']]
    df_StraightDays = df_StraightDays.drop_duplicates(['Code','Total Single Trip'])
    return df_StraightDays

#%%
    
def DaysInLastMonth_Dataframe(df_LBD,df_ExPP):
    df_LBD.sort_values(['DateVal'],ascending=[False],inplace=True)           
    df_LBD['Days_Per_Month'] = df_LBD.groupby(['Code','Period'])['Period'].transform('count')
    df_LBD.drop_duplicates(['Code'],inplace=True)   
    df_ExPP.sort_values(['DateVal'],ascending=[False],inplace=True)           
    df_ExPP['Days_Per_Month'] = df_ExPP.groupby(['Code','Period'])['Period'].transform('count')
    df_ExPP.drop_duplicates(['Code'],inplace=True)   
    MyTE_Dic = df_LBD.set_index('Code')['Days_Per_Month'].to_dict()
    Expenses_Dic = df_ExPP.set_index('Code')['Days_Per_Month'].to_dict()
    df_MyTE_Days = pd.DataFrame(MyTE_Dic.items())
    df_MyTE_Days.rename(columns={0:'Code'},inplace=True)
    df_MyTE_Days.rename(columns={1:'MyTE'},inplace=True)
    df_Expenses_Days = pd.DataFrame(Expenses_Dic.items())
    df_Expenses_Days.rename(columns={0:'Code'},inplace=True)
    df_Expenses_Days.rename(columns={1:'Expense'},inplace=True)
    df_Days = df_Expenses_Days.merge(df_MyTE_Days,on='Code')
    df_Days['Days_Per_Month'] = df_Days[["MyTE", "Expense"]].max(axis=1)
    df_Days.sort_values(['Days_Per_Month'],ascending=[False],inplace=True)  
    df_Days.drop_duplicates(['Code'],inplace=True)   
    df_Days = df_Days[['Code','Days_Per_Month']]
    return df_Days

#%%
    
def SingleTotalDays(df_StraightDays,df_LastMonth):
    df_StraightDays.rename(columns={"Total Single Trip":'Total Days in Single Trip'},inplace=True)
    return df_StraightDays

#%%
    
def Days_Info(df_LBD,df_ExPP): 
    df_LBD.rename(columns={'HomeCountry':'Home Country'},inplace=True)
    df_LBD.rename(columns={'Location1':'Destination Country/Location'},inplace=True)
    df_LBD.rename(columns={'DateVal':'Last Day in Destination Country'},inplace=True)
    df_LBD.rename(columns={'CompanyCd':'CompanyCode'},inplace=True)
    df_LBD.rename(columns={'WBS_Element':'Expense WBS'},inplace=True)
    df_LBD = df_LBD[['Code','EnterpriseId','Home Country','Destination Country/Location','Last Day in Destination Country','Days_Per_Month','Expense WBS','CompanyCode','CareerLevel']]
    df_ExPP.rename(columns={'EmployeeCountry':'Home Country'},inplace=True)
    df_ExPP.rename(columns={'CountryOfExpense':'Destination Country/Location'},inplace=True)
    df_ExPP.rename(columns={'DateVal':'Last Day in Destination Country'},inplace=True)
    df_ExPP.rename(columns={'ParentWBS':'Expense WBS'},inplace=True)
    df_ExPP.rename(columns={'Last Day in Destination Country/Location':'Last Day in Destination Country'},inplace=True)
    df_ExPP = df_ExPP[['Code','EnterpriseId','Home Country','Destination Country/Location','Last Day in Destination Country','Days_Per_Month','Expense WBS','CompanyCode','CareerLevel']]
    df_Days = pd.concat([df_LBD, df_ExPP], ignore_index=True)
    del globals()["df_LBD"]
    del globals()["df_ExPP"]
    df_Days.sort_values(['Code','Last Day in Destination Country'],ascending=[True,False],inplace=True)           
    df_Days.drop_duplicates(['Code'],inplace=True)
    return df_Days

#%%
    
def GetInfo(df_Stealth,df_Threashold,df_Days,df_StraightDays):
    df_Threashold['Stealth_ID'] = df_Threashold['Citizenship'] + df_Threashold['Host Country']
    df_Threashold.rename(columns={'Month Period':'Period'},inplace=True)
    df_Threashold.drop(['Citizenship', 'Host Country','CodeThreshold'], axis=1, inplace=True)
    df_Stealth.reset_index(inplace=True) 
    df_Stealth = df_Days.merge(df_Stealth,on='Code',how='right') 
    del globals()["df_Days"]
    df_Stealth = df_DOA[['Code','Total Days Out (MyTE vs Expense)']].merge(df_Stealth,on='Code',how='right') 
    df_Stealth['Total Days Out (MyTE vs Expense)'].fillna('N/A',inplace=True)
    del globals()["df_DOA"]
    df_Stealth = df_StraightDays.merge(df_Stealth,on='Code',how='right') 
    df_Stealth['Stealth_ID'] = df_Stealth['Home Country'] + df_Stealth['Destination Country/Location']
    df_Stealth = df_Threashold.merge(df_Stealth,on='Stealth_ID',how='right') 
    df_Stealth["Period"]= df_Stealth["Period"].str.upper().str.title() 
    del globals()["df_Threashold"]
    df_Stealth = df_TP[['Code','Travel_Plan_Number','Policy_Type','Country_Of_Citizenship','SD After Merge','ED After Merge']].merge(df_Stealth,on='Code',how='right') 
    del globals()["df_TP"]
    return df_Stealth

#%%
    
def Adding_CLandStatus(Ex_Cl,Ex_St,TE_St,df_Stealth):
    df_Stealth['Destination Market Unit'] = df_Stealth['Destination Country/Location'].map(MarketUnitDic)
    Level = {}
    inner = Ex_Cl['CareerLevel']
    Level.update(inner)
    Level.update(Te_Cl)
    df_Stealth.loc[pd.isnull(df_Stealth['CareerLevel']),'CareerLevel'] = df_Stealth['EnterpriseId'].map(Level)
    Status = {}
    Status.update(Ex_St)
    Status.update(TE_St)
    df_Stealth['Employee Status'] = df_Stealth['EnterpriseId'].map(Status)
    return df_Stealth

#%%

def SpecialThresholds(df_Stealth):
    Lev = ['Senior Manager','Managing Director','Senior Managing Director','Global Management Committee']
    Paises = ['USA','Puerto Rico']
    df_Stealth.loc[(df_Stealth['Destination Country/Location'].isin(Paises)) & (df_Stealth['CareerLevel'] == "Associate"),['Single Trip Stealth Threshold']] = 21
    df_Stealth.loc[(df_Stealth['Destination Country/Location'].isin(Paises)) & (df_Stealth['CareerLevel'] == "Associate"),['Cumulative Stealth Threshold']] = 21
    df_Stealth.loc[(df_Stealth['Destination Country/Location'].isin(Paises)) & ((df_Stealth['CareerLevel'] == "Analyst") | (df_Stealth['CareerLevel'] == "Consultant")) ,['Single Trip Stealth Threshold']] = 21
    df_Stealth.loc[(df_Stealth['Destination Country/Location'].isin(Paises)) & ((df_Stealth['CareerLevel'] == "Analyst") | (df_Stealth['CareerLevel'] == "Consultant")),['Cumulative Stealth Threshold']] = 42
    df_Stealth.loc[(df_Stealth['Destination Country/Location'].isin(Paises)) & (df_Stealth['CareerLevel'] == "Manager"),['Single Trip Stealth Threshold']] = 39
    df_Stealth.loc[(df_Stealth['Destination Country/Location'].isin(Paises)) & (df_Stealth['CareerLevel'] == "Manager"),['Cumulative Stealth Threshold']] = 59
    df_Stealth.loc[(df_Stealth['Destination Country/Location'].isin(Paises)) & (df_Stealth['CareerLevel'].isin(Lev)),['Single Trip Stealth Threshold']] = 39
    df_Stealth.loc[(df_Stealth['Destination Country/Location'].isin(Paises)) & (df_Stealth['CareerLevel'].isin(Lev)),['Cumulative Stealth Threshold']] = 89
    HomeThreshold = ['Afghanistan',	'Albania',	'Algeria',	'Andorra',	'Angola',	'Anguilla',	'Antigua and Barbuda',	'Argentina',	'Australia',	'Austria',	'Belgium',	'Brazil',	'Bulgaria',	'Canada',	'Chile',	'China/Hong Kong SAR',	'China/Mainland',	'China/Taiwan',	'Colombia',	'Cyprus',	'Czech Republic',	'Denmark',	'Dominican Rep.',	'Egypt',	'Finland',	'France',	'Germany',	'Greece',	'Hungary',	'Indonesia',	'Ireland',	'Israel',	'Italy',	'Japan',	'Jordan',	'Latvia',	'Luxembourg',	'Malaysia',	'Mauritius',	'Mexico',	'Morocco',	'Netherlands',	'New Zealand',	'Nigeria',	'Norway',	'Peru',	'Philippines',	'Poland',	'Portugal',	'Puerto Rico',	'Romania',	'Russian Federation',	'Samoa American',	'Saudi Arabia',	'Singapore',	'Slovakia',	'South Africa',	'South Korea',	'Spain',	'Sri Lanka',	'Sweden',	'Switzerland',	'Thailand',	'Turkey',	'United Arab Emirates',	'United Kingdom',	'USA',	'Venezuela']
    df_Stealth.loc[(df_Stealth['Destination Country/Location'] == "India"),['Single Trip Stealth Threshold']] = 31
    df_Stealth.loc[(df_Stealth['Destination Country/Location'] == "India") & (df_Stealth['Home Country'].isin(HomeThreshold)),['Cumulative Stealth Threshold']] = 90
    HomeThreshold = ['Canada','India','Japan','Philippines','Thailand']
    df_Stealth.loc[(df_Stealth['Destination Country/Location'] == "Sri Lanka"),['Single Trip Stealth Threshold']] = 31
    df_Stealth.loc[(df_Stealth['Destination Country/Location'] == "Sri Lanka") & (df_Stealth['Home Country'].isin(HomeThreshold)),['Cumulative Stealth Threshold']] = 31
    return df_Stealth

#%%
    
def TotalDaysColumns(df_Stealth):    
    df_Stealth['Total Cumulative Days'] = 0
    df_Stealth.loc[((df_Stealth['Period'] == "Rolling") | (df_Stealth['Period'] == "rolling")) & ((df_Stealth['Months'] == 12) | (df_Stealth['Months'] == "12")),['Total Cumulative Days']] = df_Stealth['DAYS IN LAST 12']
    df_Stealth.loc[((df_Stealth['Period'] == "Rolling") | (df_Stealth['Period'] == "rolling")) & ((df_Stealth['Months'] == 6) | (df_Stealth['Months'] == "6")) ,['Total Cumulative Days']] = df_Stealth['DAYS IN LAST 6']
    df_Stealth.loc[((df_Stealth['Period'] == "Rolling") | (df_Stealth['Period'] == "rolling")) & ((df_Stealth['Months'] == 3) | (df_Stealth['Months'] == "3")) ,['Total Cumulative Days']] = df_Stealth['DAYS IN LAST 3']
    df_Stealth.loc[((df_Stealth['Period'] == "Rolling") | (df_Stealth['Period'] == "rolling")) & ((df_Stealth['Months'] == 2) | (df_Stealth['Months'] == "2")) ,['Total Cumulative Days']] = df_Stealth['DAYS IN LAST 2']
    df_Stealth.loc[((df_Stealth['Period'] == "Calendar") | (df_Stealth['Period'] == "Calendar ")),['Total Cumulative Days']] = df_Stealth['CY']    
    df_Stealth.loc[((df_Stealth['Destination Country/Location'] == "India") | (df_Stealth['Destination Country/Location'] == "Sri Lanka")) ,['Total Cumulative Days']] = df_Stealth['FY DAYS']    
    df_Stealth.loc[((df_Stealth['Total Cumulative Days'] + 1) == df_Stealth['Total Days in Single Trip']),['Total Days in Single Trip']] = df_Stealth['Total Cumulative Days']    
    df_Stealth.loc[((df_Stealth['Total Cumulative Days'] < df_Stealth['Total Days in Single Trip'])) & (df_Stealth['Total Cumulative Days'] != 0),['Total Days in Single Trip']] = df_Stealth['Total Cumulative Days']    
    
    df_Stealth.loc[(pd.isnull(df_Stealth['Total Days in Single Trip'])),['Total Days in Single Trip']] = 0
    
    
    
    df_Stealth['Days to Reach Cumulative Threshold'] = df_Stealth['Cumulative Stealth Threshold'] - df_Stealth['Total Cumulative Days']
    df_Stealth['Days to Reach Single Trip Threshold'] = df_Stealth['Single Trip Stealth Threshold'] - df_Stealth['Total Days in Single Trip'] 
    df_Stealth['Cumulative Stealth Threshold']= pd.to_numeric(df_Stealth['Cumulative Stealth Threshold'], errors='coerce').fillna(0)
    df_Stealth['Single Trip Stealth Threshold']= pd.to_numeric(df_Stealth['Single Trip Stealth Threshold'], errors='coerce').fillna(0)
    df_Stealth['Cumulative Stealth Threshold'] = df_Stealth['Cumulative Stealth Threshold'].astype('int64')
    df_Stealth['Single Trip Stealth Threshold'] = df_Stealth['Single Trip Stealth Threshold'].astype('int64')
    #PERIODS
    df_Stealth['Months'].fillna(0,inplace=True)
    df_Stealth['Months'] = df_Stealth['Months'].astype('int64')
    df_Stealth.loc[((df_Stealth['Single Trip Stealth Threshold']) == 0),['Single Trip Stealth Threshold']]= "-"
    df_Stealth.loc[((df_Stealth['Single Trip Stealth Threshold']) == 0) & (pd.isnull(df_Stealth['Period'])==False) ,['Single Trip Stealth Threshold']]= "N/A"
    df_Stealth.loc[(df_Stealth['Single Trip Stealth Threshold'] != "N/A") & (df_Stealth['Single Trip Stealth Threshold'] != "-") ,['Single Trip Stealth Threshold']]= df_Stealth['Single Trip Stealth Threshold'].astype(str) + " (" + df_Stealth['Months'].astype(str)  + " " + df_Stealth['Period'] +  ") PP"
    df_Stealth.loc[((df_Stealth['Cumulative Stealth Threshold']==0)),['Cumulative Stealth Threshold']]= "-"
    df_Stealth.loc[(df_Stealth['Cumulative Stealth Threshold'] != "N/A") & (df_Stealth['Cumulative Stealth Threshold'] != "-"),['Cumulative Stealth Threshold']]= df_Stealth['Cumulative Stealth Threshold'].astype(str) + " (" + df_Stealth['Months'].astype(str) + " " + df_Stealth['Period'] +  ") PP"
    return df_Stealth

#%%
    
def Stealth_Status(df_Stealth):
    df_Stealth['Status - Single Trip'] = ""
    df_Stealth.loc[(df_Stealth['Days to Reach Single Trip Threshold'] <= 0),['Status - Single Trip']] = "Threshold Reached/Exceeded"
    df_Stealth.loc[(df_Stealth['Days to Reach Single Trip Threshold'] >= 1),['Status - Single Trip']] = "Within Threshold"
    df_Stealth.loc[(df_Stealth['Single Trip Stealth Threshold'] == "-"),['Status - Single Trip']] = "No Threshold available"    
    df_Stealth['Status - Cumulative'] = ""    
    df_Stealth.loc[(df_Stealth['Days to Reach Cumulative Threshold'] <= 0),['Status - Cumulative']] = "Threshold Reached/Exceeded"
    df_Stealth.loc[(df_Stealth['Days to Reach Cumulative Threshold'] >= 1),['Status - Cumulative']] = "Within Threshold"
    df_Stealth.loc[(df_Stealth['Cumulative Stealth Threshold'] == "-"),['Status - Cumulative']] = "No Threshold available"
    return df_Stealth


#%%
    
def Repeat_Offender(df_Stealth,df_Offenders):
    
    CountryPeriods = {'Andorra':'12 Rolling','Angola':'12 Rolling','Australia':'12 Rolling','Austria':'6 Rolling','Bangladesh':'12 Tax Year','Belgium':'12 Rolling','Brazil':'12 Rolling','Canada':'12 Calendar','China/Mainland':'12 Calendar','Colombia':'12 Rolling','Cyprus':'2 Rolling','Czech Republic':'12 Calendar','Denmark':'12 Rolling','Egypt':'12 Rolling','Finland':'12 Rolling','France':'3 Rolling','Germany':'6 Rolling','Ghana':'12 Rolling','Greece':'6 Rolling','China/Hong Kong SAR':'12 Rolling','Hungary':'12 Calendar','India':'12 Tax Year','Indonesia':'12 Rolling','Ireland':'12 Rolling','Israel':'12 Rolling','Italy':'6 Rolling','Japan':'12 Rolling','Kenya':'12 Rolling','Kuwait':'12 Rolling','Latvia':'12 Rolling','Luxembourg':'12 Rolling','Malaysia':'12 Rolling','Malta':'6 Rolling','Mexico':'12 Rolling','Netherlands':'6 Rolling','New Zealand':'12 Rolling','Nigeria':'12 Rolling','Norway':'12 Rolling','Oman':'12 Rolling','Philippines':'12 Calendar','Poland':'12 Rolling','Portugal':'2 Rolling','Qatar':'12 Rolling','Romania':'6 Rolling','Russian Federation':'6 Rolling','Saudi Arabia':'12 Rolling','Singapore':'12 Calendar','Slovakia':'12 Calendar','South Africa':'12 Rolling','South Korea':'12 Rolling','Spain':'2 Rolling','Sri Lanka':'12 Tax Year','Sweden':'12 Rolling','Switzerland':'12 Calendar','Thailand':'12 Rolling','Turkey':'6 Rolling','United Arab Emirates':'12 Rolling','Ukraine':'6 Rolling','United Kingdom':'12 Rolling','USA':'12 Calendar','Argentina':'12 Rolling','Chile':'12 Rolling','Bahrain':'12 Rolling'}
    df_Offenders['Period'] = df_Offenders['Host Country'].map(CountryPeriods)
    
    df_Offenders['ReportPeriod'] = pd.to_datetime(df_Offenders['ReportPeriod'])
    df_Offenders['ReportPeriod'] = df_Offenders['ReportPeriod'].apply(lambda dt: dt.replace(day=1)) 
    
    df_Offenders['Limit Date'] = ""
    df_Offenders.loc[(df_Offenders['Period'] == "2 Rolling"),['Limit Date']] = (date(int((datetime.now()).strftime("%Y")), int((datetime.now()).strftime("%m")), 1)) + dateutil.relativedelta.relativedelta(months=-2)
    df_Offenders.loc[(df_Offenders['Period'] == "3 Rolling"),['Limit Date']] = (date(int((datetime.now()).strftime("%Y")), int((datetime.now()).strftime("%m")), 1)) + dateutil.relativedelta.relativedelta(months=-3)
    df_Offenders.loc[(df_Offenders['Period'] == "6 Rolling"),['Limit Date']] = (date(int((datetime.now()).strftime("%Y")), int((datetime.now()).strftime("%m")), 1)) + dateutil.relativedelta.relativedelta(months=-6)
    df_Offenders.loc[(df_Offenders['Period'] == "12 Rolling"),['Limit Date']] = (date(int((datetime.now()).strftime("%Y")), int((datetime.now()).strftime("%m")), 1)) + dateutil.relativedelta.relativedelta(months=-12)
    df_Offenders.loc[(df_Offenders['Period'] == "12 Calendar"),['Limit Date']] = (date(int((datetime.now()).strftime("%Y")), 1, 1))
    df_Offenders.loc[(df_Offenders['Period'] == "12 Tax Year"),['Limit Date']] = (date(int((datetime.now()).strftime("%Y")), 4, 1))
 
    df_Offenders['Limit Date'] = pd.to_datetime(df_Offenders['Limit Date'])
    df_Offenders['ReportPeriod'] = pd.to_datetime(df_Offenders['ReportPeriod'])

    df_Offenders['Include?'] = True
    df_Offenders.loc[(df_Offenders['Limit Date'] >= df_Offenders['ReportPeriod']),['Include?']] = False
    df_Offenders = df_Offenders[df_Offenders['Include?'] == True]
    df_Offenders['Code'] = df_Offenders['Enterprise ID'] + df_Offenders['Host Country']
    
    RO_List = df_Offenders['Code'].drop_duplicates().tolist()
    
    df_Stealth['Repeat Offender'] = "No"
    df_Stealth.loc[(df_Stealth['Code'].isin(RO_List)) & ((df_Stealth['Status - Single Trip'] == "Threshold Reached/Exceeded") | (df_Stealth['Status - Cumulative'] == "Threshold Reached/Exceeded")) & (df_Stealth[colName[-1]] > 0),['Repeat Offender']] = "Yes"
    
    return df_Stealth

#%%

def Action_Clean_Up(df_Stealth,DOAStealth):
    df_Stealth['Action - Clean Up'] = "No action needed"
    #df_Stealth.loc[(df_Stealth['Repeat Offender'] == "No") & ((df_Stealth['Status - Single Trip'] == "Threshold Reached/Exceeded") | (df_Stealth['Status - Cumulative']== "Threshold Reached/Exceeded")),['Action - Clean Up']] = "Send Communication"
    #df_Stealth.loc[(df_Stealth['Repeat Offender'] == "Yes") & ((df_Stealth['Status - Single Trip'] == "Threshold Reached/Exceeded") | (df_Stealth['Status - Cumulative']== "Threshold Reached/Exceeded")) & (df_Stealth['ET&I Remediation Action'] == "Departure"),['Action - Clean Up']] = "Escalate"
    #df_Stealth.loc[(df_Stealth['Repeat Offender'] == "Yes") & ((df_Stealth['Status - Single Trip'] == "Threshold Reached/Exceeded") | (df_Stealth['Status - Cumulative']== "Threshold Reached/Exceeded")) & (df_Stealth['ET&I Remediation Action'] == "Registration in PMG"),['Action - Clean Up']] = "Create Travel Plan"
    #df_Stealth.loc[(df_Stealth['Repeat Offender'] == "Yes") & ((df_Stealth['Status - Single Trip'] == "Threshold Reached/Exceeded") | (df_Stealth['Status - Cumulative']== "Threshold Reached/Exceeded")) & (df_Stealth['ET&I Remediation Action'] == "Departure and Registration in PMG"),['Action - Clean Up']] = "Escalate and Create TP"
    df_Stealth.loc[(df_Stealth['Total Days Out (MyTE vs Expense)'] != "N/A") & (~df_Stealth['Code'].isin(DOAStealth)),['Action - Clean Up']] = "Days out of Assignment - Update Travel End Date"
    df_Stealth.loc[(pd.isnull(df_Stealth['Period'])),['Action - Clean Up']] = "No Threshold available, Contact ET&I Team"
    return df_Stealth

#%%
    
def CompanyName(df_Stealth):
    
    user = environ.get('USERNAME')

    Company = pd.read_excel("C:\\Users\\"+ user +"\\Desktop\Stealth\\CompanyCode.xlsx")
    Company = Company.set_index('CompanyCode')['Company Name'].to_dict()
    df_Stealth['Company Name'] = df_Stealth['CompanyCode'].map(Company)    
    return df_Stealth

#%%

def LastDayinHost(df_Stealth):
    Year = datetime.now().year
    Mes = datetime.now().month
    ReportEnd = date(Year, Mes, 1) + dateutil.relativedelta.relativedelta(days=-1)
    df_Stealth['Today'] = ReportEnd
    df_Stealth['Today'] = pd.to_datetime(df_Stealth['Today'])
    df_Stealth['Last Day of the Month in Host Country ?'] = ""
    df_Stealth.loc[(df_Stealth['Last Day in Destination Country'] == df_Stealth['Today']),['Last Day of the Month in Host Country ?']] = "Yes"
    df_Stealth.loc[(df_Stealth['Last Day of the Month in Host Country ?'] != "Yes"),['Last Day of the Month in Host Country ?']] = "No"
    return df_Stealth

#%%

def Stealth_Final_Clasification(df_Stealth):
    Quaterlys =['Bangladesh','Canada','Indonesia','Malaysia','Russian Federation','Singapore','Thailand','Ukraine']
    df_Stealth['Stealth 2'] = ""
    df_Stealth.loc[(df_Stealth['Action - Clean Up'] != "Days out of Assignment - Update Travel End Date") & (df_Stealth['Last Day of the Month in Host Country ?'] == "Yes") & ((df_Stealth['Status - Cumulative'] == "Threshold Reached/Exceeded") | (df_Stealth['Status - Single Trip'] == "Threshold Reached/Exceeded")),['Stealth 2']] = "Yes"
    df_Stealth.loc[(df_Stealth['Action - Clean Up'] != "Days out of Assignment - Update Travel End Date") & (df_Stealth['Last Day of the Month in Host Country ?'] == "Yes") & (df_Stealth['Destination Country/Location'] == "USA") & ((df_Stealth['Status - Cumulative'] == "Threshold Reached/Exceeded") | (df_Stealth['Status - Single Trip'] == "Threshold Reached/Exceeded")),['Stealth 2']] = "Yes"
    df_Stealth.loc[(df_Stealth['Stealth 2'] == ""),['Stealth 2']] = "No"
    df_Stealth['Stealth 1'] = ""
    df_Stealth.loc[(df_Stealth[colName[-1]] > 0) & (df_Stealth['Last Day of the Month in Host Country ?'] == "No") & ((df_Stealth['Status - Cumulative'] == "Threshold Reached/Exceeded") | (df_Stealth['Status - Single Trip'] == "Threshold Reached/Exceeded")) & (df_Stealth['Action - Clean Up'] != "Days out of Assignment - Update Travel End Date"),['Stealth 1']] = "Stealth Awareness"
    df_Stealth.loc[(df_Stealth[colName[-1]] > 0) & (df_Stealth['Last Day of the Month in Host Country ?'] == "No") & (df_Stealth['Destination Country/Location'] == "USA") & (df_Stealth['Total Cumulative Days'] >= 14) & (df_Stealth['Action - Clean Up'] != "Days out of Assignment - Update Travel End Date"),['Stealth 1']] = "Stealth Awareness"
    df_Stealth.loc[(df_Stealth[colName[-1]] > 0) & (df_Stealth['Destination Country/Location'] == "United Kingdom")  & ((df_Stealth['Status - Cumulative'] == "Threshold Reached/Exceeded") | (df_Stealth['Status - Single Trip'] == "Threshold Reached/Exceeded")) & (df_Stealth['Action - Clean Up'] != "Days out of Assignment - Update Travel End Date"),['Stealth 1']] = "Stealth Awareness"
    df_Stealth.loc[((df_Stealth[colName[-1]] > 0) | (df_Stealth[colName[-2]] > 0) | (df_Stealth[colName[-3]] > 0)) & (df_Stealth['Destination Country/Location'].isin(Quaterlys))  & ((df_Stealth['Status - Cumulative'] == "Threshold Reached/Exceeded") | (df_Stealth['Status - Single Trip'] == "Threshold Reached/Exceeded")) & (df_Stealth['Action - Clean Up'] != "Days out of Assignment - Update Travel End Date"),['Stealth 1']] = "Stealth Awareness"
    df_Stealth.loc[(df_Stealth['Stealth 1'] == ""),['Stealth 1']] = "No action needed"
    df_Stealth.loc[(df_Stealth['Stealth 2'] == "Yes"),['Stealth 1']] = "No action needed"
    df_Stealth['ED After Merge'] =  df_Stealth['ED After Merge'] + timedelta(days=-30)
    return df_Stealth

#%%
    
def Adding_MU(df_Stealth,MarketUnitDic):
    df_Stealth['Destination Market Unit'] = df_Stealth['Destination Country/Location'].map(MarketUnitDic)    
    return df_Stealth

#%%

def Final_Order(ColumnSource,df_Stealth):
    FirstCol = ['EnterpriseId',	'CompanyCode','Company Name','Employee Status','CareerLevel', 'Home Country',	'Destination Country/Location','Destination Market Unit','Expense WBS','Country_Of_Citizenship','Travel_Plan_Number','Policy_Type',	'SD After Merge',	'ED After Merge','Total Days Out (MyTE vs Expense)',	'Last Day in Destination Country','Last Day of the Month in Host Country ?']
    LastCol = ['Total Days in Single Trip',	'Single Trip Stealth Threshold',	'Days to Reach Single Trip Threshold',	'Status - Single Trip',	'Total Cumulative Days',	'Cumulative Stealth Threshold',	'Days to Reach Cumulative Threshold',	'Status - Cumulative',	'ET&I Remediation Action',	'Repeat Offender',	'Action - Clean Up','Stealth 1','Stealth 2']
    my_list = FirstCol + ColumnSource + LastCol
    df_Stealth = df_Stealth[my_list]
    return df_Stealth

#%%

def Detecting_VandI(df_Stealth):
    user = environ.get('USERNAME')

    df_VI = read_excel("C:\\Users\\" + user + "\\Desktop\\Stealth\\Stealth Master File.xlsx",sheet_name="ARE")
    Vi = df_VI['Company Code'].drop_duplicates().tolist()

    df_Stealth_VI = df_Stealth.copy()

    df_Stealth['Code'] = ""
    df_Stealth_VI['Code'] = ""

    df_Stealth_VI = df_Stealth_VI[df_Stealth_VI['CompanyCode'].isin(Vi)];del(df_Stealth_VI['Code'])
    df_Stealth = df_Stealth[~df_Stealth['CompanyCode'].isin(Vi)];del(df_Stealth['Code'])

    return df_Stealth,df_Stealth_VI 

#%%
    
def Detect_GMC(df_Stealth,Mds):
    df_Stealth_GMC = df_Stealth.copy()
    HighLevels = ['Senior Managing Director','Global Management Committee']
    df_Stealth = df_Stealth[(~df_Stealth['EnterpriseId'].isin(Mds))  &  (~df_Stealth['CareerLevel'].isin(HighLevels))]
    df_Stealth_GMC = df_Stealth_GMC[(df_Stealth_GMC['EnterpriseId'].isin(Mds)) | (df_Stealth_GMC['CareerLevel'].isin(HighLevels)) ]    
    return df_Stealth,df_Stealth_GMC

#%% 
    
def Export_Reports(df_Stealth,df_Stealth_VI,df_Stealth_GMC):
    user = environ.get('USERNAME')
    MonthAnalysis = datetime.now()
    MonthAnalysis = (MonthAnalysis.replace(day=1)) + dateutil.relativedelta.relativedelta(days=-1)
    MonthAnalysis = MonthAnalysis.strftime("%B")
    YearAnalysis = (datetime.now())
    YearAnalysis = (YearAnalysis.replace(day=1)) + dateutil.relativedelta.relativedelta(days=-1)
    YearAnalysis = str(YearAnalysis.strftime("%Y"))
    FinalPath = "C:\\Users\\" + user +  "\\Desktop\\Stealth\\Global Stealth Report " + MonthAnalysis + " " + str(YearAnalysis) + ".xlsx"
    with pd.ExcelWriter(FinalPath) as writer:
        df_Stealth.to_excel(writer,sheet_name='Global Stealth',index=False)
        df_Stealth_GMC.to_excel(writer,sheet_name='Global Stealth Leadership',index=False)
        df_Stealth_VI.to_excel(writer,sheet_name='V&A',index=False)
    return FinalPath

#%%
    
def Paint_Sheets(FinalPath):  
    wb = load_workbook(FinalPath)
    mysheets = ['Global Stealth','Global Stealth Leadership','V&A']  
    for lassheets in mysheets:    
        if lassheets == "V&A":
            SheetStealth = "Global Stealth V&A"
        else:
            SheetStealth = lassheets
        
        wb2 = wb.get_sheet_by_name(name=lassheets) 
        ws = wb2        
        my_blue = PatternFill(patternType='solid', fgColor=Color(rgb='00008B'))
        my_violet = PatternFill(patternType='solid', fgColor=Color(rgb='9400D3'))
        my_gray = PatternFill(patternType='solid', fgColor=Color(rgb='A9A9A9'))
        my_green = PatternFill(patternType='solid', fgColor=Color(rgb='008000'))
        my_orange = PatternFill(patternType='solid', fgColor=Color(rgb='FF8C00'))
        my_red = PatternFill(patternType='solid', fgColor=Color(rgb='B22222'))
    
        
        for i in range(1,15):
            ws.cell(row=1,column=i).fill = my_blue
            ws.cell(row=1,column=i).font = Font(color=colors.WHITE)
            ws.cell(row=1,column=i).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            
        for i in range(15,18):
            ws.cell(row=1,column=i).fill = my_orange
            ws.cell(row=1,column=i).font = Font(color=colors.WHITE)
            ws.cell(row=1,column=i).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            
        for i in range(18,42):
            ws.cell(row=1,column=i).fill = my_green
            ws.cell(row=1,column=i).font = Font(color=colors.WHITE)
            ws.cell(row=1,column=i).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            
        for i in range(42,51):
            ws.cell(row=1,column=i).fill = my_gray
            ws.cell(row=1,column=i).font = Font(color=colors.WHITE)
            ws.cell(row=1,column=i).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
               
        for i in range(51,52):
            ws.cell(row=1,column=i).fill = my_red
            ws.cell(row=1,column=i).font = Font(color=colors.WHITE)
            ws.cell(row=1,column=i).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
        for i in range(52,55):
            ws.cell(row=1,column=i).fill = my_violet
            ws.cell(row=1,column=i).font = Font(color=colors.WHITE)
            ws.cell(row=1,column=i).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                         
            
        Alfabeto = ['A','B','C','D','E','F','G','H','I','J','K','L','M','N','O','P','Q','R','S','T','U','V','W','X','Y','Z','AA','AB','AC','AD','AE','AF','AG','AH','AI','AJ','AK','AL','AM','AN','AO','AP','AQ','AR','AS','AT','AU','AV','AW','AX','AY','AZ','BA','BB']
        
        for x in Alfabeto:
            ws.column_dimensions[x].width = float(10)
#        ws.row_dimensions[1].height = float(60)
#        ws.row_dimensions[3].height = float(60)
        
        ws.insert_rows(1,3)
        ws.cell(row=2,column=2).value = SheetStealth + " " + str((date.today()).strftime("%B")) + " " + str((date.today()).strftime("%Y"))
        ws.cell(row=2,column=2).font = Font(bold=True,size=16)
    
    wb.save(FinalPath)
#%%
    
print("Importing Master File")
Mds,MarketUnitDic,ElegibleEx,df_Exceptions,df_Offenders = MasterFile()
print("Determining the report period")
Starting,Ending = ReportDates()
print("Importing SSRS Stealth Report")
df_TP = ImportStealthSSRS()
print("Importing LBD Reports")
df_LBD,df_PP,Te_Cl,TE_St = ImportLBD(ElegibleEx)
print("Importing Expense Reports")
df_EX,df_ExPP,Ex_Cl,Ex_St = ExpenseAnalysis(ElegibleEx)
print("analyzing days out of assignment in LBD")
DOA_MyTE = analisisDOA_MyTE(df_LBD,df_TP)
print("analyzing days out of assignment in Expense")
DOA_Expense = analisisDOA_Expense(df_ExPP,df_TP)
print("Combining DOA Pivots of LBD and Expenses")
df_DOA = PopulationDOA(DOA_MyTE,DOA_Expense)
print("Analyzing which are the EE with days out of Assignment considered as'Stealth'")
DOAStealth = DOA_In30(df_TP,df_LBD,df_ExPP)
print("Counting how many days per month have charged the employees")
df_EX,df_PP,colName = OrdenarColumnas(Starting,df_PP,df_EX)
print("Determining which population will be considered for the Stealth analysis")
df_Stealth,df_EX,df_PP = Selecting_Stealth_Population(df_EX,df_PP,df_TP,DOAStealth)
print("Analyzing in which pivot table the employee submitted more days (Between Work Location & Expenses)")
df_Stealth,ColumnSource = MaxandSource(Starting,df_Stealth,df_EX,df_PP)
print("Counting Rolling and Calendar days")
df_Stealth,Cols = Counting_Rolling_Calendar(df_Stealth,colName,MonthAnalysis,YearAnalysis)
print("Importing Stealth Thresholds")
df_Threashold = Importing_Thresholds()
print("Renaming Threshold columns")
df_Threashold = ConvertName(df_Threashold,'Host Country','Citizenship')
print("Analyzing which are the employees that have days straight")
df_StraightDays = StraightDays_Fist(df_ExPP,df_LBD)
df_StraightDays = Straight_days(df_StraightDays,df_Threashold,MonthAnalysis,YearAnalysis)
print("Building a dataframe with the total days per single trip")
df_StraightDays = StraightDays_Dataframe(df_StraightDays)
print("Obtaining how many days each employee has on the last Month available")
df_LastMonth = DaysInLastMonth_Dataframe(df_LBD,df_ExPP)
print("Analyzing the total days per trip")
df_StraightDays = SingleTotalDays(df_StraightDays,df_LastMonth)
print("Obtaining the general information of the Expense and Location reports")
df_Days = Days_Info(df_LBD,df_ExPP)
print("Generating the Final dataframe")
df_Stealth = GetInfo(df_Stealth,df_Threashold,df_Days,df_StraightDays)
print("Adding Employee Status and Career Level to the main dataframe")
df_Stealth = Adding_CLandStatus(Ex_Cl,Ex_St,TE_St,df_Stealth)
print("Using the special thresholds: USA & India")
df_Stealth = SpecialThresholds(df_Stealth)
print("Counting the total days for single and cummulative")
df_Stealth = TotalDaysColumns(df_Stealth)
print("Checking if the population exceeeds the Threshold")
df_Stealth = Stealth_Status(df_Stealth)
print("Checking for cases that exceeded last month.")
df_Stealth = Repeat_Offender(df_Stealth,df_Offenders)
print('Creating Action Clean Up Column')
df_Stealth = Action_Clean_Up(df_Stealth,DOAStealth)
print('Adding the Company Name')
df_Stealth = CompanyName(df_Stealth)
print('Analyzing if employees left the destination country')
df_Stealth = LastDayinHost(df_Stealth)
print("Preparing the final classification. Stealth 1 & 2")
df_Stealth = Stealth_Final_Clasification(df_Stealth)
print("Adding Market Unit Column")
df_Stealth = Adding_MU(df_Stealth,MarketUnitDic)
print("Sorting the columns for the final format")
df_Stealth = Final_Order(ColumnSource,df_Stealth)
print("Detecting V&I Population")
df_Stealth,df_Stealth_VI = Detecting_VandI(df_Stealth)
print("Detecting GMC Population")
df_Stealth,df_Stealth_GMC = Detect_GMC(df_Stealth,Mds)
print('Exporting Report to excel')
FinalPath = Export_Reports(df_Stealth,df_Stealth_VI,df_Stealth_GMC)
print('Painting and formating Final File')
Paint_Sheets(FinalPath)
print("Analysis Complete")


#
#df_LBD.drop_duplicates(['EnterpriseId'],inplace=True)   
#df_ExPP.drop_duplicates(['EnterpriseId'],inplace=True) 
#
#df_ExPP = df
#df_LBD = df
#
#
#
#Path = 'C:\\Users\\' + 'Administrator'  +  '\\Desktop\\' + 'ReporteWork.xlsx'
#
#
#with pd.ExcelWriter(Path) as writer:
#    df_LBD.to_excel(writer, sheet_name='WL', index=False)
#    df_ExPP.to_excel(writer, sheet_name='EX', index=False)
#
#    df_ExPP['CareerLevel'] = df_ExPP['EnterpriseId'].map(Ex_Cl['CareerLevel'])

